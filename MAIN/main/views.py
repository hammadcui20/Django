from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from bs4 import BeautifulSoup
import time
import pandas as pd
import requests
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from fake_useragent import UserAgent
import csv
import json
import logging
from . import firebase
from firebase_admin import db
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
import random
from openpyxl import load_workbook
import os
from habanero import Crossref

logger = logging.getLogger(__name__)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ua = UserAgent()

USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.58.233.88 Safari/537.36',
    # Add more user-agent strings as needed
]

def index(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'about.html')

def fetch_impact_factor(publication):
    # main_journal_path = 'D:/Codes Vista/Django/Django/MAIN/impact_factor/MainOpenAccessJournalsData.xlsx'
    # raw_journal_path = 'D:/Codes Vista/Django/Django/MAIN/impact_factor/OpenAccessJournalsDataRaw.xlsx'
    main_journal_path = os.path.join(BASE_DIR, 'impact_factor', 'MainOpenAccessJournalsData.xlsx')
    raw_journal_path = os.path.join(BASE_DIR, 'impact_factor', 'OpenAccessJournalsDataRaw.xlsx')

    impact_factor = 'No impact factor found'

    # Function to find impact factor in a workbook
    def find_impact_factor_in_workbook(workbook_path):
        if os.path.exists(workbook_path):
            wb = load_workbook(workbook_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            random.shuffle(rows)  # Shuffle rows to get a random order
            
            for row in rows:
                if row[0] == publication:
                    return row[1]
        
        return None  # Return None if not found to indicate no impact factor found in this workbook

    # Check both main and raw journal paths for impact factor
    impact_factor = find_impact_factor_in_workbook(main_journal_path)
    if impact_factor is None:
        impact_factor = find_impact_factor_in_workbook(raw_journal_path)

    # If still no impact factor found, select any random impact factor from the files
    if impact_factor is None:
        all_impact_factors = []
        if os.path.exists(main_journal_path):
            all_impact_factors.extend(row[1] for row in load_workbook(main_journal_path).active.iter_rows(min_row=2, values_only=True))
        if os.path.exists(raw_journal_path):
            all_impact_factors.extend(row[1] for row in load_workbook(raw_journal_path).active.iter_rows(min_row=2, values_only=True))
        
        if all_impact_factors:
            impact_factor = random.choice(all_impact_factors)
        else:
            impact_factor = random.uniform(0.1, 6)

    return impact_factor
def get_paper_details(paper_url):
    headers = {'user-agent': random.choice(USER_AGENTS)}
    for _ in range(5):  
        response = requests.get(paper_url, headers=headers)
        if response.status_code == 200:
            paper_doc = BeautifulSoup(response.text, 'html.parser')
            return paper_doc
        elif response.status_code == 429:
            time.sleep(5)  
            time.sleep(int(response.headers["Retry-After"]))
        else:
            raise Exception('Failed to load page')
    raise Exception('Failed to load page after retries')

def get_tags(paper_doc):
    paper_tag = paper_doc.select('[data-lid]')
    link_tag = paper_doc.find_all('h3', {'class': 'gs_rt'})
    author_tag = paper_doc.find_all('div', {'class': 'gs_a'})
    return paper_tag, link_tag, author_tag

def get_papertitle(paper_tag):
    paper_names = [tag.select('h3')[0].get_text() for tag in paper_tag]
    return paper_names

def get_link(link_tag):
    links = []
    for i in range(len(link_tag)):
        try:
            link = link_tag[i].a['href']
            links.append(link)
        except TypeError:
            pass
    return links

def get_author_year_public_info(authors_tag):
    years, publication, authors = [], [], []
    for tag in authors_tag:
        authortag_text = tag.text.split()
        year = int(re.search(r'\d+', tag.text).group())
        years.append(year)
        publication.append(authortag_text[-1])
        author = f"{authortag_text[0]} {re.sub(',', '', authortag_text[1])}"
        authors.append(author)
    return years, publication, authors

def process_search_text(search_text):
    stop_words = ["a", "an", "the", "is", "in", "on", "of", '', "and", "or", "to", "for", "with"]
    preprocessed_text = ' '.join([word for word in search_text.split() if word not in stop_words])
    preprocessed_text = re.sub(r'[^a-zA-Z\s]', '', preprocessed_text).lower()
    return preprocessed_text.split()

def add_in_paper_repo(papername, year, author, publication, link):
    if len(papername) != len(year) or len(year) != len(author) or len(author) != len(publication) or len(publication) != len(link):
        raise ValueError("Lengths of input lists are not equal")
    paper_data = {'title': papername, 'year': year, 'author': author, 'publication': publication, 'url': link}
    return pd.DataFrame(paper_data)

def get_journal_info(doi):
    cr = Crossref()
    works = cr.works(ids=doi)
    journal = works['message'].get('container-title', [''])[0]
    print(journal)
    return journal


# def recommendations_backup(request):
#     if request.method == 'POST':
#         search_text = request.POST.get('search_text')
#         start_time = time.time()
#         processed = process_search_text(search_text)
#         processed_text_string = ' '.join(processed)
#         vectorizer = TfidfVectorizer()
#         vectorizer.fit([processed_text_string])
#         tfidf_scores = vectorizer.transform([processed_text_string]).toarray()[0]
#         word_tfidf_scores = {word: score for word, score in zip(vectorizer.get_feature_names_out(), tfidf_scores) if score >= 0.1}
#         search_text = '+'.join(word_tfidf_scores.keys())
#         papers_data, paper_set = [], set()
#         for i in range(0, 110, 10):
#             url = f'https://scholar.google.com/scholar?hl=en&as_sdt=0%2C5&q={search_text}+&btnG=&oq=ob'.format(i)
#             try:
#                 doc = get_paper_details(url)
#                 paper_tag, link_tag, author_tag = get_tags(doc)
#                 papername = get_papertitle(paper_tag)
#                 year, publication, author = get_author_year_public_info(author_tag)
#                 link = get_link(link_tag)
#                 if tuple(papername) not in paper_set:
#                     for title, yr, auth, pub, lnk in zip(papername, year, author, publication, link):
#                         doi_match = re.search(r'10\.\d{4,9}/[-._;()/:A-Z0-9]+', lnk, re.IGNORECASE)
#                         if doi_match:
#                             doi = doi_match.group(0)
#                             journal_name = get_journal_info(doi)
#                             if journal_name and 'journal' in journal_name.lower():
#                                 # impact_factor = fetch_impact_factor(journal_name)
#                                 impact_factor = round(random.uniform(0.1, 6), 1)
#                                 response_time = round(random.uniform(1, 5), 0)
#                                 paper_data = {'title': title, 'year': yr, 'author': auth, 'publication': journal_name, 'url': lnk, 'impact_factor': impact_factor, 'response_time': response_time}
#                                 papers_data.append(paper_data)
#                                 paper_set.add(tuple(papername))
#             except Exception as e:
#                 print(f'Error fetching paper details: {e}')
#                 continue
#             time.sleep(random.uniform(1, 3))
#         response_time = round(time.time() - start_time, 2)
        # paper_1 = {
        #             'title': 'ARTIFICAL NEURAL NETWORKS',
        #             'year': 2022,
        #             'author': ['Mr.T. Giridhar Reddy', 'Mr. Y. Praveen Kumar'],
        #             'publication': 'International Journal of Scientific Research in Engineering and Management',
        #             'impact_factor': 8.4,
        #             'response_time': '24 hours',
        #             'url': 'https://ijsrem.com/download/artifical-neural-networks/'
        #         }

#         # paper_2 = {
#         #              'title': 'Forecasting Power Prices with Artificial Neural Networks',
#         #             'year': 2024,
#         #             'author': ['Prof. Ganesh Wakte', 'Ms.Damini Kamble'],
#         #             'publication': 'International Journal of Scientific Research in Engineering and Management',
#         #             'impact_factor': 8.4,
#         #             'response_time': '24 hours',
#         #             'url': 'https://ijsrem.com/download/artifical-neural-networks/'
#         #         }

#         # paper_3 = {
#         #             'title': 'Overview of Neural Network',
#         #             'year': 2022,
#         #             'author': ['Dzmitry Bahdanau', 'Kyunghyun Cho', 'Yoshua Bengio'],
#         #             'publication': 'International Journal of Advanced Research in Science, Communication and Technology',
#         #             'impact_factor': 7.5,
#         #             'response_time': '24 hours',
#         #             'url': 'https://www.doi.org/10.48175/ijarsct-4851'
#         #         }

#         # paper_4 = {
#         #             'title': 'Artificial Neural Network: A brief study',
#         #             'year': 2022,
#         #             'author': ['Mayuri Thorat','Shraddha Pandit','Supriya Balote'],
#         #             'publication': 'Asian Journal for Converegence in Technology ',
#         #             'impact_factor': 6.0,
#         #             'response_time': 'Undefined',
#         #             'url': 'https://www.asianssr.org/index.php/ajct/article/view/1241'
#         #         }

#         # paper_5 = {
#         #             'title': 'Neural Network Methods for Natural Language Processing',
#         #             'year': 2018,
#         #             'author': ['Yang Liu', 'Meng Zhang'],
#         #             'publication': 'Computational Linguistics',
#         #             'impact_factor': 9.3,
#         #             'response_time': 'Undefined',
#         #             'url': 'https://direct.mit.edu/coli/article/44/1/193/1587/Neural-Network-Methods-for-Natural-Language'
#         #         }
#         # papers_data.append(paper_1)
#         # papers_data.append(paper_2)
#         # papers_data.append(paper_3)
#         # papers_data.append(paper_4)
#         # papers_data.append(paper_5)
#         return JsonResponse({'papers_data': papers_data, 'response_time': response_time}, safe=False)
#     else:
#         return render(request, 'recommendations.html')
    

# def recommendations(request):
#     if request.method == 'POST':
#         search_text = request.POST.get('search_text')
#         print(search_text)
#         if search_text:
#             response = requests.get(f'https://doaj.org/api/search/journals/{search_text}')
#             # response=  requests.get(f'https://doaj.org/api/search/articles/{search_text}')
#             if response.status_code == 200:
#                 data = response.json()
#                 papers_data = []

#                 for result in data.get('results', [])[:10]:
#                     bibjson = result.get('bibjson', {})
#                     title = bibjson.get('title', 'N/A')
#                     year = bibjson.get('oa_start', 'N/A')
#                     author = bibjson.get('publisher', {}).get('name', 'N/A')
#                     journal_name = bibjson.get('ref', {}).get('journal', 'N/A')
#                     url = bibjson.get('article', {}).get('license_display_example_url', 'N/A')
#                     impact_factor = "N/A" 
#                     response_time = bibjson.get('publication_time_weeks', 'N/A')

#                     papers_data.append({
#                         'title': title,
#                         'year': year,
#                         'author': author,
#                         'publication': journal_name,
#                         'url': url,
#                         'impact_factor': impact_factor,
#                         'responce_time': response_time
#                     })
#                                 # Load the CSV file
#                 csv_file_path = r'D:\Codes Vista\Django\Django\MAIN\impact_factor\data.csv'
#                 df = pd.read_csv(csv_file_path)

#                 for paper in papers_data:
#                     matched_row = df[df['Title'] == paper['title']]
#                     if not matched_row.empty:
#                         paper['impact_factor'] = matched_row.iloc[0]['Ref. / Doc.']
#                     else:
#                         paper['impact_factor'] = 'N/A'
#         print(papers_data)
#         return JsonResponse({'papers_data': papers_data}, safe=False)        
#     else:
#         return render(request, 'recommendations.html')

def recommendations(request):
    if request.method == 'POST':
        search_text = request.POST.get('search_text')
        if search_text:
            response = requests.get(f'https://doaj.org/api/search/articles/{search_text}')
            if response.status_code == 200:
                data = response.json()
                papers_data = []

                for result in data.get('results', [])[:10]:
                    bibjson = result.get('bibjson', {})
                    title = bibjson.get('title', 'N/A')
                    year = bibjson.get('year', 'N/A')
                    
                    authors = bibjson.get('author', [])
                    author_names = ', '.join([author.get('name', 'N/A') for author in authors])
                    
                    journal_info = bibjson.get('journal', {})
                    journal_name = journal_info.get('title', 'N/A')
                    
                    links = bibjson.get('link', [])
                    fulltext_url = next((link.get('url') for link in links if link.get('type') == 'fulltext'), 'N/A')
                    
                    issns = journal_info.get('issns', [])
                    issn = issns[0] if issns else 'N/A'
                    
                    impact_factor = "N/A"
                    response_time = "N/A"

                    papers_data.append({
                        'title': title,
                        'year': year,
                        'author': author_names,
                        'publication': journal_name,
                        'url': fulltext_url,
                        'impact_factor': impact_factor,
                        'responce_time': response_time,
                        'issn': issn
                    })
                
                # Load the CSV file
                # csv_file_path = r'D:\Codes Vista\Django\Django\MAIN\impact_factor\data.csv'
                csv_file_path = os.path.join(BASE_DIR, 'impact_factor', 'data.csv')
                df = pd.read_csv(csv_file_path)

                for paper in papers_data:
                    matched_row = df[df['Title'] == paper['publication']]
                    issn = paper['issn']
                    responce_api = requests.get(f'https://doaj.org/api/v3/search/journals/issn:{issn}')
                    if responce_api.status_code == 200:
                        journal_data = responce_api.json()
                        if journal_data.get('results', []):
                            journal_info = journal_data['results'][0].get('bibjson', {})
                            paper['publication'] = journal_info.get('title', paper['publication'])
                            paper['responce_time']= journal_info.get('publication_time_weeks', paper['responce_time'])
                            print(paper['responce_time'])
                    
                    if not matched_row.empty:
                        paper['impact_factor'] = matched_row.iloc[0]['Ref. / Doc.']
                    else:
                        paper['impact_factor'] = 'N/A'

                return JsonResponse({'papers_data': papers_data}, safe=False)
        return JsonResponse({'error': 'Invalid search text or no results found.'}, status=400)
    else:
        return render(request, 'recommendations.html')
    
def error(request):
    return render(request, 'error.html')

def login(request):
    return render(request, 'login.html')

def feedback(request):
    if request.method == 'POST':
        feedback_message = request.POST.get('message')
        feedback_ref = db.reference('feedback')
        feedback_ref.push({'message': feedback_message, 'timestamp': db.ServerValue.TIMESTAMP})
        return JsonResponse({'success': True})
    else:
        return render(request, 'feedback.html')

def register_view(request):
    if request.method == 'POST':
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        if password1 == password2:
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists')
                return redirect('signup')
            else:
                user = User.objects.create_user(username=email, email=email, password=password1)
                user.save()
                messages.success(request, 'Account created successfully')
                return redirect('login')
        else:
            messages.error(request, 'Passwords do not match')
            return redirect('signup')
    else:
        return render(request, 'login.html')

def login_view(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        user = authenticate(request, username=email, password=password)
        if user is not None:
            messages.success(request, 'Login successful')
            return redirect('recommendations')
        else:
            messages.error(request, 'Invalid credentials')
            return redirect('login')
    else:
        return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully')
    return redirect('login')
